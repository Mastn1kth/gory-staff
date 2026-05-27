from __future__ import annotations

import argparse
import datetime as dt
import subprocess
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data" / "Gory-Data.xlsx"

STAFF_ROLES = {"pending", "manager", "administrator", "hostess", "waiter", "kitchen", "bar", "technical", "technician"}
STAFF_STATUSES = {"on_shift", "off_shift", "sick", "vacation", "inactive", "blocked", "fired"}
GUEST_STATUSES = {"active", "blocked", "inactive"}
LOYALTY_LEVELS = {"bronze", "silver", "gold", "platinum"}

HEADER_ALIASES = {
    "id": "id",
    "имя": "name",
    "name": "name",
    "телефон": "phone",
    "phone": "phone",
    "логин": "login",
    "login": "login",
    "пароль": "password",
    "password": "password",
    "новый пароль": "new_password",
    "new_password": "new_password",
    "роль": "role",
    "role": "role",
    "должность": "position",
    "position": "position",
    "статус": "status",
    "status": "status",
    "комментарий": "comment",
    "comment": "comment",
    "создан": "created_at",
    "created_at": "created_at",
    "дата рождения": "birthday",
    "birthday": "birthday",
    "email": "email",
    "бонусы": "bonus_balance",
    "bonus_balance": "bonus_balance",
    "уровень лояльности": "loyalty_level",
    "loyalty_level": "loyalty_level",
    "номер карты": "card_number",
    "card_number": "card_number",
    "визиты": "visits_count",
    "visits_count": "visits_count",
    "потрачено": "total_spent",
    "total_spent": "total_spent",
    "средний чек": "average_check",
    "average_check": "average_check",
    "согласие на маркетинг": "marketing_consent",
    "marketing_consent": "marketing_consent",
    "согласие на персональные данные": "personal_data_consent",
    "personal_data_consent": "personal_data_consent",
}

ROLE_ALIASES = {
    "ожидает роли": "pending",
    "новый сотрудник": "pending",
    "управляющий": "manager",
    "администратор": "administrator",
    "хостес": "hostess",
    "официант": "waiter",
    "кухня": "kitchen",
    "повар": "kitchen",
    "бар": "bar",
    "бармен": "bar",
    "технический сотрудник": "technical",
    "технический": "technical",
    "охрана": "technical",
    "техник системы": "technician",
    "техник": "technician",
    "тех администратор": "technician",
    "технический администратор": "technician",
}

STAFF_STATUS_ALIASES = {
    "на смене": "on_shift",
    "не на смене": "off_shift",
    "больничный": "sick",
    "отпуск": "vacation",
    "неактивен": "inactive",
    "заблокирован": "blocked",
    "уволен": "fired",
}

GUEST_STATUS_ALIASES = {
    "активен": "active",
    "заблокирован": "blocked",
    "неактивен": "inactive",
}

LOYALTY_ALIASES = {
    "бронза": "bronze",
    "серебро": "silver",
    "золото": "gold",
    "платина": "platinum",
}


def compact(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def normalize_header(value: Any) -> str:
    text = compact(value).lower()
    return HEADER_ALIASES.get(text, text)


def normalize_phone(value: Any) -> str:
    raw = compact(value)
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 10:
        return f"+7{digits}"
    if len(digits) == 11 and digits.startswith("8"):
        return f"+7{digits[1:]}"
    if len(digits) == 11 and digits.startswith("7"):
        return f"+{digits}"
    if raw.startswith("+") and 10 <= len(digits) <= 15:
        return f"+{digits}"
    return raw


def normalize_bool(value: Any) -> bool:
    text = compact(value).lower()
    return text in {"1", "true", "yes", "y", "да", "истина", "on"}


def normalize_int(value: Any, default: int = 0) -> int:
    text = compact(value).replace(" ", "").replace(",", ".")
    if not text:
        return default
    return int(float(text))


def normalize_date(value: Any) -> str | None:
    if value is None or compact(value) == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = compact(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"Bad date value: {text}")


def sql(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    text = str(value).replace("'", "''")
    return f"'{text}'"


def safe_choice(value: Any, allowed: set[str], default: str, aliases: dict[str, str] | None = None) -> str:
    text = compact(value).lower()
    if aliases:
        text = aliases.get(text, text)
    return text if text in allowed else default


def find_sheet(wb, required_headers: set[str]):
    for ws in wb.worksheets:
        headers = {normalize_header(cell.value) for cell in ws[1]}
        if required_headers.issubset(headers):
            return ws
    return None


def read_rows(ws) -> list[dict[str, Any]]:
    headers = [normalize_header(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(compact(value) == "" for value in values):
            continue
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
        rows.append(row)
    return rows


def read_workbook(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    staff_sheet = find_sheet(wb, {"name", "login", "role"})
    guest_sheet = find_sheet(wb, {"name", "phone", "bonus_balance"})
    return (
        read_rows(staff_sheet) if staff_sheet else [],
        read_rows(guest_sheet) if guest_sheet else [],
    )


def bcrypt_hash(password: str) -> str:
    result = subprocess.run(
        [
            "node",
            "-e",
            "const bcrypt=require('bcryptjs'); console.log(bcrypt.hashSync(process.argv[1], 10));",
            password,
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "Could not hash password with bcryptjs")
    return result.stdout.strip()


def referral_code() -> str:
    return f"GOR{uuid.uuid4().hex[:8].upper()}"


def card_number() -> str:
    return f"GC{uuid.uuid4().hex[:10].upper()}"


def staff_statement(row: dict[str, Any]) -> str | None:
    name = compact(row.get("name"))
    login = compact(row.get("login")).lower()
    if not name and not login:
        return None
    if not name or not login:
        raise ValueError("Staff row needs both name and login.")

    password = compact(row.get("new_password") or row.get("password"))
    password_hash = None
    if password:
        if len(password) < 8:
            raise ValueError(f"Password for staff login {login} must be at least 8 characters.")
        password_hash = bcrypt_hash(password)
    if not compact(row.get("id")) and password_hash is None:
        raise ValueError(f"New staff login {login} needs new_password.")
    password_plain = password if password_hash else None

    user_id = compact(row.get("id")) or str(uuid.uuid4())
    phone = compact(row.get("phone"))
    role = safe_choice(row.get("role"), STAFF_ROLES, "waiter", ROLE_ALIASES)
    position = compact(row.get("position")) or "Сотрудник"
    status = safe_choice(row.get("status"), STAFF_STATUSES, "off_shift", STAFF_STATUS_ALIASES)
    comment = compact(row.get("comment"))

    return f"""
DO $$
DECLARE target_id TEXT;
BEGIN
  SELECT id INTO target_id
  FROM users
  WHERE id = {sql(user_id)} OR login = {sql(login)}
  LIMIT 1;

  IF target_id IS NULL THEN
    INSERT INTO users (id, name, phone, login, password_hash, password_plain, role, position, status, photo_url, comment, created_at)
    VALUES ({sql(user_id)}, {sql(name)}, {sql(phone)}, {sql(login)}, {sql(password_hash)}, {sql(password_plain)}, {sql(role)}, {sql(position)}, {sql(status)}, NULL, {sql(comment)}, NOW());
  ELSE
    UPDATE users
    SET name = {sql(name)},
        phone = {sql(phone)},
        login = {sql(login)},
        password_hash = COALESCE({sql(password_hash)}, password_hash),
        password_plain = COALESCE({sql(password_plain)}, password_plain),
        role = {sql(role)},
        position = {sql(position)},
        status = {sql(status)},
        comment = {sql(comment)}
    WHERE id = target_id;
  END IF;
END $$;
"""


def guest_statement(row: dict[str, Any]) -> str | None:
    name = compact(row.get("name"))
    phone = normalize_phone(row.get("phone"))
    if not name and not phone:
        return None
    if not name or not phone:
        raise ValueError("Guest row needs both name and phone.")

    guest_id = compact(row.get("id")) or str(uuid.uuid4())
    birthday = normalize_date(row.get("birthday"))
    email = compact(row.get("email")) or None
    bonus_balance = max(0, normalize_int(row.get("bonus_balance")))
    loyalty_level = safe_choice(row.get("loyalty_level"), LOYALTY_LEVELS, "bronze", LOYALTY_ALIASES)
    status = safe_choice(row.get("status"), GUEST_STATUSES, "active", GUEST_STATUS_ALIASES)
    visits_count = max(0, normalize_int(row.get("visits_count")))
    total_spent = max(0, normalize_int(row.get("total_spent")))
    average_check = max(0, normalize_int(row.get("average_check")))
    marketing_consent = normalize_bool(row.get("marketing_consent"))
    personal_data_consent = normalize_bool(row.get("personal_data_consent"))
    guest_card = compact(row.get("card_number")) or card_number()
    code = referral_code()

    return f"""
DO $$
DECLARE target_id TEXT;
BEGIN
  SELECT id INTO target_id
  FROM guest_users
  WHERE id = {sql(guest_id)} OR phone = {sql(phone)}
  LIMIT 1;

  IF target_id IS NULL THEN
    target_id := {sql(guest_id)};
    INSERT INTO guest_users
      (id, name, phone, birthday, email, bonus_balance, loyalty_level, referral_code, status, visits_count, total_spent, average_check, marketing_consent, personal_data_consent, created_at, updated_at)
    VALUES
      (target_id, {sql(name)}, {sql(phone)}, {sql(birthday)}, {sql(email)}, {bonus_balance}, {sql(loyalty_level)}, {sql(code)}, {sql(status)}, {visits_count}, {total_spent}, {average_check}, {sql(marketing_consent)}, {sql(personal_data_consent)}, NOW(), NOW());
  ELSE
    UPDATE guest_users
    SET name = {sql(name)},
        phone = {sql(phone)},
        birthday = {sql(birthday)},
        email = {sql(email)},
        bonus_balance = {bonus_balance},
        loyalty_level = {sql(loyalty_level)},
        status = {sql(status)},
        visits_count = {visits_count},
        total_spent = {total_spent},
        average_check = {average_check},
        marketing_consent = {sql(marketing_consent)},
        personal_data_consent = {sql(personal_data_consent)},
        updated_at = NOW()
    WHERE id = target_id;
  END IF;

  INSERT INTO guest_cards (id, guest_id, card_number, level, issued_at, status, created_at, updated_at)
  VALUES ({sql(str(uuid.uuid4()))}, target_id, {sql(guest_card)}, {sql(loyalty_level)}, NOW(), 'active', NOW(), NOW())
  ON CONFLICT (guest_id)
  DO UPDATE SET card_number = EXCLUDED.card_number, level = EXCLUDED.level, status = 'active', updated_at = NOW();
END $$;
"""


def build_import_sql(staff_rows: list[dict[str, Any]], guest_rows: list[dict[str, Any]]) -> tuple[str, int, int]:
    statements = ["BEGIN;"]
    staff_count = 0
    guest_count = 0

    for row in staff_rows:
        statement = staff_statement(row)
        if statement:
            statements.append(statement)
            staff_count += 1

    for row in guest_rows:
        statement = guest_statement(row)
        if statement:
            statements.append(statement)
            guest_count += 1

    statements.append("COMMIT;")
    return "\n".join(statements), staff_count, guest_count


def run_psql(script: str) -> None:
    command = ["docker", "exec", "-i", "gory-staff-postgres", "psql", "-U", "gory", "-d", "gory_staff", "-v", "ON_ERROR_STOP=1"]
    result = subprocess.run(command, cwd=ROOT, input=script, text=True, encoding="utf-8", capture_output=True)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Excel import failed")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import staff and guests from data/Gory-Data.xlsx")
    parser.add_argument("path", nargs="?", default=str(DEFAULT_INPUT))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    path = Path(args.path)
    if not path.is_absolute():
        path = ROOT / path
    if not path.exists():
        raise SystemExit(f"Excel file not found: {path}")

    staff_rows, guest_rows = read_workbook(path)
    script, staff_count, guest_count = build_import_sql(staff_rows, guest_rows)
    if args.dry_run:
        print(f"DRY RUN: staff rows={staff_count}, guest rows={guest_count}")
        return

    run_psql(script)
    print(f"Imported staff rows: {staff_count}")
    print(f"Imported guest rows: {guest_count}")


if __name__ == "__main__":
    main()
