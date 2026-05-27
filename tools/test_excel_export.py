from __future__ import annotations

import importlib.util
import unittest

from openpyxl import Workbook

from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "tools" / "export_excel_tables.py"
spec = importlib.util.spec_from_file_location("export_excel_tables", SCRIPT)
excel_export = importlib.util.module_from_spec(spec)
assert spec.loader is not None
spec.loader.exec_module(excel_export)


class ExcelExportTests(unittest.TestCase):
    def test_staff_sheet_uses_russian_headers_and_visible_password_column(self):
        wb = Workbook()
        wb.remove(wb.active)
        rows = [
            {
                "id": "u-test",
                "name": "Test Staff",
                "phone": "+7 900 000-00-01",
                "login": "worker@example.test",
                "password": "StaffTestPass-2026!",
                "new_password": "",
                "role": "waiter",
                "position": "Waiter",
                "status": "off_shift",
                "comment": "",
                "created_at": "",
            }
        ]

        excel_export.add_sheet(wb, excel_export.SHEET_STAFF, rows, excel_export.STAFF_COLUMNS)
        ws = wb[excel_export.SHEET_STAFF]
        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]

        self.assertEqual(ws.title, "Сотрудники")
        self.assertIn("Пароль", headers)
        self.assertIn("Новый пароль", headers)
        self.assertNotIn("new_password", headers)
        self.assertEqual(values[headers.index("Пароль")], "StaffTestPass-2026!")
        self.assertEqual(values[headers.index("Роль")], "Официант")
        self.assertEqual(values[headers.index("Статус")], "Не на смене")

    def test_guest_sheet_uses_russian_headers(self):
        wb = Workbook()
        wb.remove(wb.active)

        excel_export.add_sheet(wb, excel_export.SHEET_GUESTS, [], excel_export.GUEST_COLUMNS)
        headers = [cell.value for cell in wb[excel_export.SHEET_GUESTS][1]]

        self.assertIn("Имя", headers)
        self.assertIn("Телефон", headers)
        self.assertIn("Бонусы", headers)
        self.assertNotIn("bonus_balance", headers)


if __name__ == "__main__":
    unittest.main()
