from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "tools" / "import_excel_tables.py"
spec = importlib.util.spec_from_file_location("import_excel_tables", SCRIPT)
excel_import = importlib.util.module_from_spec(spec)
assert spec.loader is not None
spec.loader.exec_module(excel_import)


class ExcelImportTests(unittest.TestCase):
    def make_workbook(self) -> Path:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        path = Path(tmp.name)

        wb = Workbook()
        info = wb.active
        info.title = "Info"
        info.append(["ignore"])

        staff = wb.create_sheet("Сотрудники")
        staff.append(["ID", "Имя", "Телефон", "Логин", "Пароль", "Новый пароль", "Роль", "Должность", "Статус", "Комментарий"])
        staff.append(["u-test", "Test Staff", "+7 900 000-00-01", "worker@example.test", "", "", "Официант", "Waiter", "Не на смене", ""])

        guests = wb.create_sheet("Гости")
        guests.append(["ID", "Имя", "Телефон", "Дата рождения", "Email", "Бонусы", "Уровень лояльности", "Статус", "Визиты", "Потрачено", "Средний чек"])
        guests.append(["", "Test Guest", "8 900 000-00-02", "1990-01-02", "guest@example.test", 500, "silver", "active", 2, 3000, 1500])
        wb.save(path)
        return path

    def test_reads_staff_and_guests_from_workbook(self):
        path = self.make_workbook()
        try:
            staff_rows, guest_rows = excel_import.read_workbook(path)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(len(staff_rows), 1)
        self.assertEqual(len(guest_rows), 1)
        self.assertEqual(staff_rows[0]["login"], "worker@example.test")
        self.assertEqual(staff_rows[0]["role"], "Официант")
        self.assertEqual(excel_import.normalize_phone(guest_rows[0]["phone"]), "+79000000002")

    def test_builds_import_sql_for_existing_staff_and_new_guest(self):
        path = self.make_workbook()
        try:
            staff_rows, guest_rows = excel_import.read_workbook(path)
            script, staff_count, guest_count = excel_import.build_import_sql(staff_rows, guest_rows)
        finally:
            path.unlink(missing_ok=True)

        self.assertEqual(staff_count, 1)
        self.assertEqual(guest_count, 1)
        self.assertIn("UPDATE users", script)
        self.assertIn("password_plain", script)
        self.assertIn("'waiter'", script)
        self.assertIn("'off_shift'", script)
        self.assertIn("INSERT INTO guest_users", script)
        self.assertIn("+79000000002", script)


if __name__ == "__main__":
    unittest.main()
