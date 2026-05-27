from __future__ import annotations

import csv
import subprocess
from io import StringIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "Gory-Data.xlsx"

SHEET_INFO = "Инструкция"
SHEET_STAFF = "Сотрудники"
SHEET_GUESTS = "Гости"

STAFF_COLUMNS = [
    ("id", "ID"),
    ("name", "Имя"),
    ("phone", "Телефон"),
    ("login", "Логин"),
    ("password", "Пароль"),
    ("new_password", "Новый пароль"),
    ("role", "Роль"),
    ("position", "Должность"),
    ("status", "Статус"),
    ("comment", "Комментарий"),
    ("created_at", "Создан"),
]

GUEST_COLUMNS = [
    ("id", "ID"),
    ("name", "Имя"),
    ("phone", "Телефон"),
    ("birthday", "Дата рождения"),
    ("email", "Email"),
    ("bonus_balance", "Бонусы"),
    ("loyalty_level", "Уровень лояльности"),
    ("card_number", "Номер карты"),
    ("status", "Статус"),
    ("visits_count", "Визиты"),
    ("total_spent", "Потрачено"),
    ("average_check", "Средний чек"),
    ("marketing_consent", "Согласие на маркетинг"),
    ("personal_data_consent", "Согласие на персональные данные"),
    ("created_at", "Создан"),
]

STAFF_HEADERS = [title for _, title in STAFF_COLUMNS]
GUEST_HEADERS = [title for _, title in GUEST_COLUMNS]

ROLE_LABELS = {
    "pending": "Ожидает роли",
    "manager": "Управляющий",
    "administrator": "Администратор",
    "hostess": "Хостес",
    "waiter": "Официант",
    "kitchen": "Кухня",
    "bar": "Бар",
    "technical": "Технический сотрудник",
    "technician": "Техник системы",
}

STAFF_STATUS_LABELS = {
    "on_shift": "На смене",
    "off_shift": "Не на смене",
    "sick": "Больничный",
    "vacation": "Отпуск",
    "inactive": "Неактивен",
    "blocked": "Заблокирован",
    "fired": "Уволен",
}

GUEST_STATUS_LABELS = {
    "active": "Активен",
    "blocked": "Заблокирован",
    "inactive": "Неактивен",
}

LOYALTY_LABELS = {
    "bronze": "Бронза",
    "silver": "Серебро",
    "gold": "Золото",
    "platinum": "Платина",
}

STAFF_QUERY = """
SELECT
  id,
  name,
  phone,
  login,
  COALESCE(password_plain, '') AS password,
  '' AS new_password,
  role,
  position,
  status,
  comment,
  created_at
FROM users
ORDER BY role, name
"""

GUESTS_QUERY = """
SELECT
  gu.id,
  gu.name,
  gu.phone,
  gu.birthday,
  gu.email,
  gu.bonus_balance,
  gu.loyalty_level,
  gc.card_number,
  gu.status,
  gu.visits_count,
  gu.total_spent,
  gu.average_check,
  gu.marketing_consent,
  gu.personal_data_consent,
  gu.created_at
FROM guest_users gu
LEFT JOIN guest_cards gc ON gc.guest_id = gu.id AND gc.status = 'active'
WHERE gu.deleted_at IS NULL
ORDER BY gu.created_at DESC
"""

STAFF_FALLBACK = [
    {
        "id": "u-admin",
        "name": "Георгий Казбеков",
        "phone": "+7 900 100-10-01",
        "login": "owner@gory.local",
        "password": "",
        "new_password": "",
        "role": "manager",
        "position": "Управляющий",
        "status": "on_shift",
        "comment": "Отвечает за зал, финансы, персонал и развитие ресторана.",
        "created_at": "",
    },
    {
        "id": "u-hostess",
        "name": "Мария Лазарева",
        "phone": "+7 900 100-10-03",
        "login": "hostess",
        "password": "",
        "new_password": "",
        "role": "hostess",
        "position": "Хостес",
        "status": "on_shift",
        "comment": "Встреча гостей, рассадка, звонки по броням.",
        "created_at": "",
    },
]


def read_query(query: str) -> list[dict[str, str]]:
    command = [
        "docker",
        "exec",
        "gory-staff-postgres",
        "psql",
        "-U",
        "gory",
        "-d",
        "gory_staff",
        "-c",
        f"COPY ({query}) TO STDOUT WITH CSV HEADER",
    ]
    result = subprocess.run(command, cwd=ROOT, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if result.returncode != 0 or not result.stdout.strip():
        raise RuntimeError(result.stderr.strip() or "PostgreSQL is not available")
    return list(csv.DictReader(StringIO(result.stdout)))


def display_value(row: dict[str, str], key: str) -> str:
    value = "" if row.get(key) is None else str(row.get(key, ""))
    if key == "role":
        return ROLE_LABELS.get(value, value)
    if key == "status":
        return STAFF_STATUS_LABELS.get(value, GUEST_STATUS_LABELS.get(value, value))
    if key == "loyalty_level":
        return LOYALTY_LABELS.get(value, value)
    if key in {"marketing_consent", "personal_data_consent"}:
        return "Да" if value.lower() in {"true", "1", "yes", "да"} else "Нет"
    return value


def fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 44))
        ws.column_dimensions[column_letter].width = max(12, max_length + 2)


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, str]], columns: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title)
    ws.append([header for _, header in columns])
    for row in rows:
        ws.append([display_value(row, key) for key, _ in columns])

    header_fill = PatternFill("solid", fgColor="7A2638")
    header_font = Font(color="FFF8EA", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row, 1)}"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    fit_columns(ws)


def add_info_sheet(wb: Workbook) -> None:
    info = wb.active
    info.title = SHEET_INFO
    rows = [
        ["Горы", "Excel-таблица сотрудников и гостей"],
        ["Обновление", 'Кнопка "Открыть Excel" и START_GORY_STAFF.bat выгружают свежие данные из PostgreSQL.'],
        ["Импорт", 'После ручных правок закройте Excel и нажмите "Импорт Excel" в панели управления.'],
        [
            SHEET_STAFF,
            "Новый сотрудник: заполните Имя, Логин и Новый пароль. Колонка Пароль показывает последний пароль, заданный через приложение или импорт. Старые пароли из хэша восстановить нельзя.",
        ],
        [SHEET_GUESTS, "Новый гость: заполните Имя и Телефон. Бонусы и статусы можно править вручную."],
        ["Файл", str(OUTPUT)],
    ]
    for row in rows:
        info.append(row)
    info["A1"].font = Font(bold=True, size=16, color="7A2638")
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    fit_columns(info)


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    try:
        staff_rows = read_query(STAFF_QUERY)
    except Exception:
        staff_rows = STAFF_FALLBACK

    try:
        guest_rows = read_query(GUESTS_QUERY)
    except Exception:
        guest_rows = []

    wb = Workbook()
    add_info_sheet(wb)
    add_sheet(wb, SHEET_STAFF, staff_rows, STAFF_COLUMNS)
    add_sheet(wb, SHEET_GUESTS, guest_rows, GUEST_COLUMNS)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
